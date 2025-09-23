# src/osm_robot/robot.py
from __future__ import annotations
from typing import List, Dict, Any, Tuple
from openpyxl import load_workbook

def read_excel_region(path: str, sheet: str, a1_range: str) -> List[Dict[str, Any]]:
    """
    Prečíta A1-oblasť z Excelu a vráti zoznam buniek (row, col, cell, dtype, value_or_fx).
    - dtype: 'blank' | 'text' | 'number' | 'formula' | 'other'
    - value_or_fx: hodnota (text/číslo) alebo originál vzorec pri formula
    """
    r1, c1, r2, c2 = _parse_range(a1_range)
    wb = load_workbook(filename=path, data_only=False, read_only=False)
    ws = wb[sheet]

    out: List[Dict[str, Any]] = []
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            v = cell.value
            if isinstance(v, str) and v.startswith("="):
                dtype = "formula"
                val = v
            elif isinstance(v, str):
                dtype = "text"
                val = v
            elif v is None:
                dtype = "blank"
                val = ""
            elif isinstance(v, (int, float)):
                dtype = "number"
                val = v
            else:
                dtype = "other"
                val = v
            out.append({
                "row": r,
                "col": c,
                "cell": f"{_col_letter(c)}{r}",
                "dtype": dtype,
                "value_or_fx": val,
            })
    return out

# -- pomocné funkcie --
def _parse_range(a1: str) -> Tuple[int, int, int, int]:
    import re
    m = re.match(r'^\s*([A-Za-z]+)(\d+)\s*:\s*([A-Za-z]+)(\d+)\s*$', a1)
    if not m:
        raise ValueError(f"Invalid A1 range: {a1}")
    c1 = _col_index(m.group(1)); r1 = int(m.group(2))
    c2 = _col_index(m.group(3)); r2 = int(m.group(4))
    if r2 < r1 or c2 < c1:
        raise ValueError("Range end must be below/right of start")
    return r1, c1, r2, c2

def _col_index(col: str) -> int:
    col = col.upper()
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - 64)
    return n

def _col_letter(idx: int) -> str:
    out = []
    while idx:
        idx, r = divmod(idx - 1, 26)
        out.append(chr(r + 65))
    return "".join(reversed(out))


if __name__ == "__main__":
    import argparse, csv, os
    ap = argparse.ArgumentParser()
    ap.add_argument("--file", "-f", required=True)
    ap.add_argument("--sheet", "-s", required=True)
    ap.add_argument("--range", "-r", required=True)
    ap.add_argument("--out", "-o", default="data_out/cells.csv")
    args = ap.parse_args()

    os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
    cells = read_excel_region(args.file, args.sheet, args.range)

    with open(args.out, "w", newline="", encoding="utf-8") as fh:
        w = csv.DictWriter(fh, fieldnames=["row","col","cell","dtype","value_or_fx"])
        w.writeheader()
        for c in cells:
            w.writerow({k: c[k] for k in ["row","col","cell","dtype","value_or_fx"]})

    print(f"Wrote {args.out} ({len(cells)} cells)")
