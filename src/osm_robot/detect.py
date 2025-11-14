import re
from typing import Any, Dict

import yaml
from openpyxl import load_workbook

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTHS_FULL = [
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December",
]
FULL2SHORT = {f.lower(): s for f, s in zip(MONTHS_FULL, MONTHS)}


def _col_to_idx(col: str) -> int:
    col = col.upper()
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - 64)
    return n


def _idx_to_col(i: int) -> str:
    out = []
    while i:
        i, r = divmod(i - 1, 26)
        out.append(chr(r + 65))
    return "".join(reversed(out))


def _parse_range(a1: str):
    m = re.match(r"^\s*([A-Za-z]+)(\d+)\s*:\s*([A-Za-z]+)(\d+)\s*$", a1)
    if not m:
        raise ValueError(f"Bad A1: {a1}")
    c1, r1, c2, r2 = (
        _col_to_idx(m.group(1)),
        int(m.group(2)),
        _col_to_idx(m.group(3)),
        int(m.group(4)),
    )
    if r2 < r1 or c2 < c1:
        raise ValueError("Range end must be below/right of start")
    return r1, c1, r2, c2


def _norm_header(h: str) -> str:
    if h is None:
        return ""
    x = str(h).strip()
    if not x:
        return ""
    x = re.sub(r"\s+", "", x)
    xl = x.lower()
    if xl in FULL2SHORT:
        return FULL2SHORT[xl]
    if re.fullmatch(r"\d{1,2}", x):
        return "W" + x
    m = re.match(r"^[Ww]0?(\d{1,2})$", x)
    if m:
        return "W" + str(int(m.group(1)))
    return x


def detect_block(path: str, sheet: str, a1_range: str) -> Dict[str, Any]:
    r1, c1, r2, c2 = _parse_range(a1_range)
    wb = load_workbook(path, data_only=False, read_only=False)
    ws = wb[sheet]

    headers = []
    for c in range(c1 + 1, c2 + 1):
        headers.append(_norm_header(ws.cell(row=r1, column=c).value))

    week_pat = re.compile(r"^W(\d{1,2})$")
    if headers and all(week_pat.match(h) for h in headers if h):
        layout = "weeks"
        cols_by = "week"
    elif headers and all((h[:3].title() in MONTHS) for h in headers if h):
        layout = "months"
        cols_by = "month"
    else:
        common = {"YTD", "Budget", "Var", "Var%", "LE", "LEVar", "LEVar%", "Actual", "Plan"}
        hs = [h for h in headers if h]
        if hs and set(hs).issubset({*common}):
            layout = "static_cols"
            cols_by = hs
        else:
            layout = "unknown"
            cols_by = hs

    row_labels = []
    for r in range(r1 + 1, r2 + 1):
        v = ws.cell(row=r, column=c1).value
        row_labels.append("" if v is None else str(v).strip())
    while row_labels and row_labels[-1] == "":
        row_labels.pop()

    return {
        "sheet": sheet,
        "anchor": f"{_idx_to_col(c1 + 1)}{r1 + 1}",
        "layout": layout,
        "rows_by": "metric",
        "row_map": row_labels,
        "cols_by": cols_by,
        "source_view": "TBD_view",
        "filters": "country='AT', year=2025",
        "value": "SUM(value)",
        "rounding": 0,
        "format": "#,##0;[Red]-#,##0",
        "tolerance": 0.5,
    }


if __name__ == "__main__":
    import argparse
    import os

    ap = argparse.ArgumentParser()
    ap.add_argument("--file", "-f", required=True)
    ap.add_argument("--sheet", "-s", required=True)
    ap.add_argument("--range", "-r", required=True)
    ap.add_argument("--out", "-o", default="data_out/block.yaml")
    args = ap.parse_args()

    os.makedirs(os.path.dirname(args.out) or ".", exist_ok=True)
    spec = detect_block(args.file, args.sheet, args.range)

    with open(args.out, "w", encoding="utf-8") as fh:
        yaml.safe_dump(spec, fh, sort_keys=False, allow_unicode=True)
    print(f"Detected layout={spec['layout']} → wrote {args.out}")
