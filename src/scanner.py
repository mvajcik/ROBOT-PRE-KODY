# src/scanner.py
from __future__ import annotations
import argparse
import csv
import os
import re
from typing import Iterable, Tuple, Optional

from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple, get_column_letter

HEADER = ["sheet", "row", "col", "value", "formula"]

# ---------------------------
# A1 helpers
# ---------------------------
def _parse_a1(a1: str) -> Tuple[int, int]:
    r, c = coordinate_to_tuple(a1)  # (row, col_idx)
    return r, c

def _clean_part(part: str) -> str:
    """Remove sheet qualifier, $, and spaces."""
    s = part.strip()
    if "!" in s:  # 'Sheet'!A1 → A1
        s = s.split("!", 1)[1]
    if len(s) >= 2 and s[0] == "'" and s[-1] == "'":
        s = s[1:-1]
    return s.replace("$", "").replace(" ", "")

def _to_float(x) -> Optional[float]:
    if x is None or x == "":
        return None
    try:
        return float(str(x).replace(" ", "").replace(",", "."))
    except ValueError:
        return None

def _sum_range(ws_values, a1_range: str) -> float:
    """Sum a single A1 range (cleaned)."""
    if ":" not in a1_range:
        a1_range = f"{a1_range}:{a1_range}"
    tl, br = a1_range.split(":")
    r1, c1 = _parse_a1(tl)
    r2, c2 = _parse_a1(br)
    total = 0.0
    for row in ws_values.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2, values_only=True):
        for v in row:
            fv = _to_float(v)
            if fv is not None:
                total += fv
    return total

def _value_cell(ws_values, a1: str) -> float:
    r, c = _parse_a1(a1)
    return _to_float(ws_values.cell(row=r, column=c).value) or 0.0

# ---------------------------
# Lightweight formula eval
# ---------------------------
_SUM_FUNS = {
    "SUM", "SUMA", "SUMME", "SÚČET", "SUCET",
}

# SUBTOTAL local names
_SUBTOTAL_FUNS = {
    "SUBTOTAL",         # EN
    "MEDZISÚČET", "MEDZISUCET",  # SK
    "TEILERGEBNIS",     # DE
    "PODSOUČET", "PODSOUCET",    # CZ
    "PODSUMOWANIE",     # PL
}

def _eval_sum_like(ws_values, formula: str) -> Optional[float]:
    """=SUM(...)/local variants, supports ',' or ';' and multiple parts."""
    m = re.fullmatch(r"=\s*([A-ZÁÄČĎÉÍĹĽŇÓÔŔŠŤÚÝŽ]+)\s*\((.+)\)\s*", formula, re.IGNORECASE)
    if not m:
        return None
    fn = m.group(1).upper()
    if fn not in _SUM_FUNS:
        return None
    inner = m.group(2).replace(";", ",")
    parts = [p for p in (x.strip() for x in inner.split(",")) if p]
    total = 0.0
    for p in parts:
        rng = _clean_part(p)
        total += _sum_range(ws_values, rng)
    return total

def _eval_subtotal(ws_values, formula: str) -> Optional[float]:
    """
    Evaluate SUBTOTAL-like functions:
      SUBTOTAL(fn_num, range[, range2...])
      Only fn_num 9 or 109 (SUM) are supported.
    """
    m = re.fullmatch(r"=\s*([A-ZÁÄČĎÉÍĹĽŇÓÔŔŠŤÚÝŽ]+)\s*\((.+)\)\s*", formula, re.IGNORECASE)
    if not m:
        return None
    fn = m.group(1).upper()
    if fn not in _SUBTOTAL_FUNS:
        return None

    inner = m.group(2).replace(";", ",")
    # first token is function number
    tokens = [t.strip() for t in inner.split(",") if t.strip()]
    if not tokens:
        return None
    try:
        fn_num = int(tokens[0])
    except ValueError:
        return None
    if fn_num not in (9, 109):
        return None

    ranges = tokens[1:]
    if not ranges:
        return 0.0
    total = 0.0
    for p in ranges:
        rng = _clean_part(p)
        total += _sum_range(ws_values, rng)
    return total

def _eval_plus_chain(ws_values, formula: str) -> Optional[float]:
    """=A1+B2+... only '+' operator, single-cell refs."""
    if not formula.startswith("="):
        return None
    expr = formula[1:].strip()
    if any(ch in expr for ch in "():,;*/-"):
        return None
    parts = [p for p in (x.strip() for x in expr.split("+")) if p]
    if not parts:
        return None
    total = 0.0
    for p in parts:
        ref = _clean_part(p)
        if ":" in ref:
            return None
        total += _value_cell(ws_values, ref)
    return total

# ---------------------------
# Merged cells
# ---------------------------
def _resolve_merged_anchor(ws, row: int, col: int) -> Optional[Tuple[int, int]]:
    coord = f"{get_column_letter(col)}{row}"
    for rng in ws.merged_cells.ranges:
        if coord in rng:
            if row != rng.min_row or col != rng.min_col:
                return rng.min_row, rng.min_col
            return None
    return None

def _read_formula_and_value(ws_formula, ws_values, row: int, col: int) -> Tuple[str, object]:
    cell_f = ws_formula.cell(row=row, column=col)
    cell_v = ws_values.cell(row=row, column=col)

    raw = cell_f.value
    val = cell_v.value

    # merged fallback
    if (raw is None or raw == "") and (val is None or val == ""):
        anchor = _resolve_merged_anchor(ws_formula, row, col)
        if anchor:
            ar, ac = anchor
            cell_f = ws_formula.cell(row=ar, column=ac)
            cell_v = ws_values.cell(row=ar, column=ac)
            raw = cell_f.value
            val = cell_v.value

    formula = raw if isinstance(raw, str) and raw.startswith("=") else ""

    # compute if cached missing
    if (val is None or val == "") and formula:
        for evaluator in (_eval_sum_like, _eval_subtotal, _eval_plus_chain):
            v = evaluator(ws_values, formula)
            if v is not None:
                val = v
                break

    value = "" if val is None else val
    return formula, value

# ---------------------------
# Helper: ensure value
# ---------------------------
def _ensure_value_filled(value, formula):
    return formula if (value is None or str(value) == '') and (formula not in (None, '')) else value

# ---------------------------
# Emit
# ---------------------------
def _emit_block(ws_formula, ws_values, top_left: str, bottom_right: str, sheet_name: str, writer: csv.writer) -> None:
    r1, c1 = coordinate_to_tuple(top_left)
    r2, c2 = coordinate_to_tuple(bottom_right)
    for r_off in range(0, r2 - r1 + 1):
        for c_off in range(0, c2 - c1 + 1):
            abs_row = r1 + r_off
            abs_col = c1 + c_off
            col_letter = get_column_letter(abs_col)
            formula, value = _read_formula_and_value(ws_formula, ws_values, abs_row, abs_col)
            # poistka pred zápisom
            value = _ensure_value_filled(value, formula)
            writer.writerow([sheet_name, abs_row, col_letter, value, formula])

# ---------------------------
# Public API
# ---------------------------
def scan_block(xlsx_path: str, sheet_name: str, range_str: str, out_csv: str) -> None:
    try:
        top_left, bottom_right = range_str.split(":")
    except ValueError:
        raise ValueError(f"Range musí byť vo forme 'B2:E10', dostal som: {range_str}")

    wb_f = load_workbook(xlsx_path, data_only=False, read_only=False)   # formulas
    wb_v = load_workbook(xlsx_path, data_only=True,  read_only=False)   # cached
    if sheet_name not in wb_f.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb_f.sheetnames}")
    ws_f = wb_f[sheet_name]
    ws_v = wb_v[sheet_name]

    os.makedirs(os.path.dirname(out_csv) or ".", exist_ok=True)
    with open(out_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(HEADER)
        _emit_block(ws_f, ws_v, top_left, bottom_right, sheet_name, w)
    print(f"Wrote {out_csv}")

def scan_blocks(xlsx_path: str, sheet_name: str, ranges: Iterable[str], out_csv: str, append: bool = False) -> None:
    wb_f = load_workbook(xlsx_path, data_only=False, read_only=False)
    wb_v = load_workbook(xlsx_path, data_only=True,  read_only=False)
    if sheet_name not in wb_f.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found. Available: {wb_f.sheetnames}")
    ws_f = wb_f[sheet_name]
    ws_v = wb_v[sheet_name]

    os.makedirs(os.path.dirname(out_csv) or ".", exist_ok=True)
    mode = "a" if append and os.path.exists(out_csv) else "w"
    with open(out_csv, mode, newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        if mode == "w":
            w.writerow(HEADER)
        for r in ranges:
            try:
                tl, br = r.split(":")
            except ValueError:
                raise ValueError(f"Range musí byť vo forme 'B2:E10', dostal som: {r}")
            _emit_block(ws_f, ws_v, tl, br, sheet_name, w)
    print(f"Wrote {out_csv}")

# ---------------------------
# CLI
# ---------------------------
def main() -> None:
    p = argparse.ArgumentParser(description="Scan Excel cell block(s) to CSV (cell map).")
    p.add_argument("--file", required=True, help="Path to .xlsx file")
    p.add_argument("--sheet", required=True, help="Worksheet name")
    p.add_argument("--range", dest="ranges", action="append", required=True,
                   help="A1 range, e.g. B2:E10. Repeatable: --range B2:E6 --range I6:I6")
    p.add_argument("--out", default="data_out/scan.csv", help="Output CSV path")
    p.add_argument("--append", action="store_true", help="Append to existing CSV (keep header only once)")
    args = p.parse_args()

    if len(args.ranges) == 1 and not args.append:
        scan_block(args.file, args.sheet, args.ranges[0], args.out)
    else:
        scan_blocks(args.file, args.sheet, args.ranges, args.out, append=args.append)

if __name__ == "__main__":
    main()

# --- Legacy API (for tests) ---
def scan_block_legacy(file, *, sheet, top_left, bottom_right):
    wb = load_workbook(filename=str(file), data_only=False, read_only=False)
    ws = wb[sheet]
    r1, c1 = coordinate_to_tuple(top_left)
    r2, c2 = coordinate_to_tuple(bottom_right)
    out = []
    for row in ws.iter_rows(min_row=r1, max_row=r2, min_col=c1, max_col=c2, values_only=False):
        for cell in row:
            out.append({
                "row": cell.row,
                "col": cell.column,
                "address": cell.coordinate,
                "value": cell.value,
            })
    return out  # ← toto bolo chýbajúce (a omylom bolo v helperi)
