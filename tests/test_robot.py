# tests/test_robot.py
from pathlib import Path
from openpyxl import Workbook
from src.osm_robot.robot import read_excel_region

def _make_tmp_excel(tmp_path: Path) -> str:
    """Vytvorí malý Excel s číslami, textom a vzorcom."""
    wb = Workbook()
    ws = wb.active
    ws.title = "AT_YTD"
    ws["B2"] = "Hello"        # text
    ws["C2"] = 1234           # number
    ws["D2"] = "=C2*2"        # formula
    ws["E2"] = None           # blank
    path = tmp_path / "mini.xlsx"
    wb.save(path)
    return str(path)

def test_read_excel_region(tmp_path: Path):
    xlsx = _make_tmp_excel(tmp_path)
    cells = read_excel_region(xlsx, "AT_YTD", "B2:E2")
    # očakávame 4 bunky
    assert len(cells) == 4
    # typy buniek
    assert cells[0]["dtype"] == "text" and cells[0]["value_or_fx"] == "Hello"
    assert cells[1]["dtype"] == "number" and cells[1]["value_or_fx"] == 1234
    assert cells[2]["dtype"] == "formula" and str(cells[2]["value_or_fx"]).startswith("=C2*2")
    assert cells[3]["dtype"] == "blank" and cells[3]["value_or_fx"] == ""
