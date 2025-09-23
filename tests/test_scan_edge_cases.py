# tests/test_scan_edge_cases.py
from pathlib import Path
import csv
from openpyxl import Workbook
from src.scanner import scan_block

def read_csv(path: Path):
    with path.open(newline="", encoding="utf-8") as f:
        return list(csv.reader(f))

def test_empty_cells_and_formula(tmp_path):
    # 1) pripravíme dočasné xlsx
    xlsx = tmp_path / "edge.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "SHEET1"

    # range B2:E6
    ws["B2"] = 123                 # číslo
    ws["C2"] = None                # prázdna bunka
    ws["D2"] = "=B2*2"             # formula
    ws["E6"] = "text"              # text
    wb.save(xlsx)

    # 2) spustíme scanner
    out_csv = tmp_path / "scan.csv"
    scan_block(str(xlsx), "SHEET1", "B2:E6", str(out_csv))

    rows = read_csv(out_csv)
    header, data = rows[0], rows[1:]
    assert header == ["sheet", "row", "col", "value", "formula"]

    # 3) nájdeme konkrétne bunky
    as_dicts = [
        {"sheet": r[0], "row": int(r[1]), "col": r[2], "value": r[3], "formula": r[4]}
        for r in data
    ]

    b2 = next(r for r in as_dicts if r["row"] == 2 and r["col"] == "B")
    c2 = next(r for r in as_dicts if r["row"] == 2 and r["col"] == "C")
    d2 = next(r for r in as_dicts if r["row"] == 2 and r["col"] == "D")
    e6 = next(r for r in as_dicts if r["row"] == 6 and r["col"] == "E")

    assert b2["value"] == "123" or b2["value"] == 123    # toleruj číselný/str výstup
    assert c2["value"] == ""                             # prázdna bunka → ""
    assert d2["formula"].startswith("=") and d2["formula"] == "=B2*2"
    assert e6["value"] == "text"
