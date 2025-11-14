# tests/test_detect.py
from pathlib import Path

from openpyxl import Workbook

from src.osm_robot.detect import detect_block


def _wb_weeks(tmp_path: Path) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "AT_YTD"
    ws["B2"] = "W1"
    ws["C2"] = "W02"
    ws["D2"] = "w3"
    ws["E2"] = "W4"
    ws["B3"] = "Turnover"
    ws["C3"] = 100
    ws["D3"] = 200
    ws["E3"] = 300
    p = tmp_path / "weeks.xlsx"
    wb.save(p)
    return str(p)


def test_detect_block_weeks(tmp_path: Path):
    xlsx = _wb_weeks(tmp_path)
    spec = detect_block(xlsx, "AT_YTD", "B2:E10")
    assert spec["layout"] == "weeks"
    assert spec["cols_by"] == "week"
    assert spec["rows_by"] == "metric"
    assert spec["row_map"][0] == "Turnover"
